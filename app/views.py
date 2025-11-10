from django.db import transaction
from django.utils.crypto import get_random_string
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.decorators import api_view
from rest_framework import generics
from django.shortcuts import get_object_or_404
from django.db import models 
from django.shortcuts import render
from rest_framework.decorators import action,permission_classes
from .permissions import IsAdministrador, IsPropietario
from rest_framework.parsers import JSONParser
from rest_framework import viewsets,status
from  django.utils import timezone
from django.core.exceptions import PermissionDenied
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.views import TokenObtainPairView
from django.contrib.auth.models import Group, Permission as AuthPermission
from rest_framework import serializers
import logging
from django.conf import settings
import stripe 
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status

stripe.api_key = settings.STRIPE_SECRET_KEY
# views.py
from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from .models import Carrito, CarritoItem
from .serializers import CarritoSerializer, CarritoItemSerializer
from django.shortcuts import get_object_or_404
from django.shortcuts import get_object_or_404
from django.utils.crypto import get_random_string
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Carrito, CarritoItem, Producto,Venta
from .serializers import CarritoSerializer, VentaSerializer,CarritoItemSerializer
#pago
from django.contrib.contenttypes.fields import GenericForeignKey
from django.contrib.contenttypes.models import ContentType
import stripe
from django.db.models import Prefetch
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated


logger = logging.getLogger(__name__)

from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.utils import timezone
import logging
from rest_framework.response import Response
from rest_framework import status





from .models import (
    Rol, Usuario,Bitacora,DetalleBitacora,Pago,
    NotificacionPush,DispositivoMovil)
from  .mixin import BitacoraLoggerMixin
from .serializers import (ChangePasswordSerializer,PagoSerializer,
    RolSerializer,UsuarioSerializer, UsuarioMeSerializer,
    LogoutSerializer,MyTokenPairSerializer,GroupSerializer,
    DispositivoMovilSerializer,NotificacionPushSerializer,BitacoraSerializer,DetalleBitacoraSerializer
   
)
class MyTokenObtainPairView(TokenObtainPairView): 
    serializer_class = MyTokenPairSerializer
    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.user  # ← ESTE es el usuario autenticado

        # IP (X-Forwarded-For si hay proxy; si no, REMOTE_ADDR)
        xff = request.META.get('HTTP_X_FORWARDED_FOR')
        ip = (xff.split(',')[0].strip() if xff else request.META.get('REMOTE_ADDR')) or None

        # User-Agent como "device" (o None si vacío)
        device = request.META.get('HTTP_USER_AGENT') or None

        # Registrar login en bitácora
        Bitacora.objects.create(
            usuario=user,
            login=timezone.now(),
            ip=ip,
            device=device
        )
        logger.info('el usuario ingreso al perfil',)

        return Response(serializer.validated_data, status=status.HTTP_200_OK)

from rest_framework.permissions import IsAuthenticated



class RolViewSet(BitacoraLoggerMixin,viewsets.ModelViewSet):
    
    queryset = Rol.objects.all()
    serializer_class = RolSerializer
    # Puedes añadir permisos aquí si necesitas restringir el acceso
    # permission_classes = [IsAuthenticated]
    def create(self, request, *args, **kwargs):
        # Muestra los datos que llegan a la vista
        print("Datos de la solicitud (request.data):", request.data)
        
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Muestra los datos validados antes de la creación
        print("Datos validados por el serializer:", serializer.validated_data)
        
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

class UsuarioViewSet(BitacoraLoggerMixin, viewsets.ModelViewSet):
    """
    ViewSet para manejar usuarios.
    Incluye un endpoint para enviar correos electrónicos.
    """
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer
    filter_backends = [DjangoFilterBackend]  # Agregar soporte para filtrado
    filterset_fields = ['rol__nombre']  # Filtrado por nombre de rol
    permission_classes = [IsAuthenticated]  # Seguridad adecuada, solo usuarios autenticados

    @action(detail=False, methods=['post'], url_path='enviar-correo', permission_classes=[IsAuthenticated])
    def enviar_correo(self, request):
        """
        Endpoint para enviar un correo electrónico.
        """
        subject = request.data.get('subject')
        message = request.data.get('message')

        if not subject or not message:
            return Response(
                {"error": "Subject and message are required."},
                status=status.HTTP_400_BAD_REQUEST
            )
        

    def get_queryset(self):
        # Obtener el queryset base
        queryset = super().get_queryset()

        # Si el usuario autenticado es un Administrador
        # y no está pidiendo el endpoint 'me' (que maneja su propio serializador)
        # excluye al propio usuario de la lista.
        # Es importante que el usuario autenticado no sea excluido cuando accede a '/me'
        # o cuando un administrador necesita ver su propio perfil en otra parte.

        if self.request.user.is_authenticated:
            # Comprobamos si el usuario tiene un rol asignado y si ese rol es 'Administrador'
            # Asumiendo que el campo 'rol' es un ForeignKey al modelo Rol y tiene un campo 'nombre'
            is_admin = hasattr(self.request.user, 'rol') and \
                       self.request.user.rol and \
                       self.request.user.rol.nombre == 'Administrador'

            # Además, podemos verificar si la acción actual NO es 'retrieve' o 'me'
            # para evitar excluir al administrador de su propio detalle o del endpoint 'me'.
            # La acción 'list' es donde generalmente querrías esta exclusión.
            if is_admin and self.action == 'list':
                queryset = queryset.exclude(id=self.request.user.id)
            
            # Si se está filtrando por rol, también aplicamos ese filtro
            # filterset_fields ya maneja esto, pero si lo haces manualmente...
            # rol_nombre_param = self.request.query_params.get('rol__nombre')
            # if rol_nombre_param:
            #     queryset = queryset.filter(rol__nombre=rol_nombre_param)

        return queryset
   
    @action(
        detail=False,
        methods=['get'],
        url_path='me',
        permission_classes=[IsAuthenticated]
    )
    def me(self, request):
        """
        Endpoint exclusivo para el usuario autenticado.
        """
        # Usa el nuevo serializador específico en lugar del serializador por defecto del ViewSet
        serializer = UsuarioMeSerializer(request.user)
        return Response(serializer.data)
    @action(
        detail=True,
        methods=['post'],
        url_path='cambiar-contrasena',  # Cambiar la URL también si lo deseas
        permission_classes=[IsAuthenticated]
    )
    def cambiar_contrasena(self, request, pk=None):
        print("📥 Payload recibido en backend (cambiar_contrasena):", request.data)

        target_user = get_object_or_404(Usuario, pk=pk)  # Obtener el usuario objetivo

    # --- Lógica de Permisos de Vista ---
        is_request_user_admin = (request.user.is_superuser or
                             (hasattr(request.user, 'rol') and request.user.rol and request.user.rol.nombre == 'Administrador'))

        if not is_request_user_admin:  # Si el usuario no es admin/superuser
            if target_user.id != request.user.id:
                return Response(
                    {"detail": "No tienes permiso para cambiar la contraseña de otro usuario."},
                    status=status.HTTP_403_FORBIDDEN
                )
    
    # Lógica para la validación de la nueva contraseña con el serializer
        ser = ChangePasswordSerializer(
            data=request.data,
            context={"request": request, "user": target_user}
        )
        ser.is_valid(raise_exception=True)

    # Si pasa validación, se setea la nueva contraseña
        new_pwd = ser.validated_data["new_password"]
        target_user.set_password(new_pwd)
        target_user.save(update_fields=["password"])

    # (Opcional) Registrar en bitácora esta acción específica
        try:
        # Aquí iría tu lógica para registrar en la bitácora si la tienes implementada
            pass
        except Exception:
            pass

    # Responder con un estado 204 No Content
        return Response(status=status.HTTP_204_NO_CONTENT)


    @action(detail=False, methods=['get'], url_path='clientes', url_name='clientes')
    def clientes(self, request):                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                             
        clientes=self.queryset.filter(rol__nombre="Cliente")
        serializer=self.get_serializer(clientes,many=True)
        return Response(serializer.data)
    @action(
        detail=True,
        methods=['get'],
        url_path='perfil',
        permission_classes=[IsAuthenticated]
    )
    def perfil(self, request, pk=None):
        """
        Endpoint para obtener el perfil del usuario.
        """
        # Obtener el usuario por ID (pk)
        usuario = get_object_or_404(Usuario, pk=pk)

        # Usamos el serializador para devolver la información del perfil
        serializer = UsuarioSerializer(usuario)

        return Response(serializer.data)







class MyTokenObtainPairView(TokenObtainPairView): 
    serializer_class = MyTokenPairSerializer
    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.user  # ← ESTE es el usuario autenticado

        # IP (X-Forwarded-For si hay proxy; si no, REMOTE_ADDR)
        xff = request.META.get('HTTP_X_FORWARDED_FOR')
        ip = (xff.split(',')[0].strip() if xff else request.META.get('REMOTE_ADDR')) or None

        # User-Agent como "device" (o None si vacío)
        device = request.META.get('HTTP_USER_AGENT') or None

        # Registrar login en bitácora
        Bitacora.objects.create(
            usuario=user,
            login=timezone.now(),
            ip=ip,
            device=device
        )
        logger.info('el usuario ingreso al perfil',)

        return Response(serializer.validated_data, status=status.HTTP_200_OK)
#vistas
"""
{
  "refresh": "...",
  "password": "..."
}

"""
class LogoutView(APIView):
    """
    Endpoint de **logout**.
    Requiere `{"refresh": "<jwt-refresh-token>"}` en el cuerpo (JSON).
    Blacklistea el refresh token mediante SimpleJWT y registra el logout en Bitacora si corresponde.
    Retorna 204 en caso de éxito.
    """
    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser]  # fuerza a intentar parsear JSON

    def post(self, request):
        # --- DEBUG: cuerpo crudo + datos parseados + headers ---
        raw = request.body.decode("utf-8", errors="replace")
        headers = {
            k: v for k, v in request.META.items()
            if k.startswith("HTTP_") or k in ("CONTENT_TYPE", "CONTENT_LENGTH")
        }

        #logger.info("=== RAW BODY === %s", raw)
        #logger.info("=== PARSED DATA === %s", request.data)
        #logger.info("=== HEADERS === %s", headers)
    
        # invalidamos el refresh token
        serializer = LogoutSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        # ——————— Registro de logout ———————
        bit = Bitacora.objects.filter(
            usuario=request.user,
            logout__isnull=True
        ).last()
        if bit:
            print('no se esta cerrando seccion ')
            bit.logout = timezone.now()
            bit.save()

        return Response(status=status.HTTP_204_NO_CONTENT)
class AuthPermissionSerializer(serializers.ModelSerializer):
    class Meta:
        model = AuthPermission
        fields = ['id', 'codename', 'name']  
class AuthPermissionViewSet(BitacoraLoggerMixin,viewsets.ReadOnlyModelViewSet):
    """
    Lista todas las Django Permissions.
    """
    queryset = AuthPermission.objects.all()
    serializer_class = AuthPermissionSerializer
    permission_classes = [IsAdministrador]  
class GroupViewSet(BitacoraLoggerMixin,viewsets.ModelViewSet):

    queryset = Group.objects.all()
    serializer_class = GroupSerializer

class DispositivoMovilViewSet(BitacoraLoggerMixin,viewsets.ModelViewSet):
    queryset = DispositivoMovil.objects.all()
    serializer_class = DispositivoMovilSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend] 
    filterset_fields = ['activo', 'usuario', 'token_fcm'] # <-- ¡AQUÍ ESTÁ LA CLAVE!
    search_fields = ['modelo_dispositivo', 'sistema_operativo', 'token_fcm']
    ordering_fields = ['fecha_registro', 'ultima_conexion']


from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from .fcm_service import enviar_notificacion_fcm
from .models import Usuario, DispositivoMovil
from rest_framework.permissions import IsAuthenticated

   

class NotificacionPushViewSet(BitacoraLoggerMixin, viewsets.ModelViewSet):
    queryset = NotificacionPush.objects.all()
    serializer_class = NotificacionPushSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['estado', 'tipo', 'usuario']
    search_fields = ['titulo', 'cuerpo']
    ordering_fields = ['fecha_envio', 'fecha_lectura']
    
    @action(detail=False, methods=['post'], url_path='enviar-notificacion')
    def enviar_notificacion(self, request):
        """
        Enviar una notificación push a un dispositivo específico.
        """
        # Obtenemos los datos enviados en el request
        dispositivo_id = request.data.get('dispositivo_id')  # Cambiamos usuario_id a dispositivo_id
        titulo = request.data.get('titulo')
        cuerpo = request.data.get('cuerpo')
        tipo = request.data.get('tipo', 'sistema')
        datos_adicionales = request.data.get('datos_adicionales', {})

        # Validación de parámetros
        if not dispositivo_id or not titulo or not cuerpo:
            return Response({"detail": "Faltan parámetros necesarios: dispositivo_id, titulo, cuerpo."}, status=status.HTTP_400_BAD_REQUEST)

        # Obtener el dispositivo y verificar si existe
        try:
            dispositivo = DispositivoMovil.objects.get(id=dispositivo_id)
        except DispositivoMovil.DoesNotExist:
            return Response({"detail": "Dispositivo no encontrado."}, status=status.HTTP_404_NOT_FOUND)

        # Enviar la notificación
        try:
            # Llamamos a la función que maneja el envío de la notificación
            notificaciones_enviadas = enviar_notificacion_fcm(
                usuario=dispositivo.usuario,  # Usamos el usuario del dispositivo
                titulo=titulo,
                cuerpo=cuerpo,
                tipo=tipo,
                datos_adicionales=datos_adicionales
            )

            # Responder con las notificaciones enviadas
            if notificaciones_enviadas:
                return Response({
                    "detail": "Notificación enviada exitosamente.",
                    "notificaciones": NotificacionPushSerializer(notificaciones_enviadas, many=True).data
                }, status=status.HTTP_200_OK)
            else:
                return Response({"detail": "No se enviaron notificaciones."}, status=status.HTTP_400_BAD_REQUEST)
        
        except Exception as e:
            return Response({"detail": f"Error enviando notificación: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class BitacoraViewSet(BitacoraLoggerMixin,viewsets.ModelViewSet):
    queryset=Bitacora.objects.all()
    serializer_class=BitacoraSerializer
    permission_classes=[IsAuthenticated]

class DetalleBitacoraViewSet(BitacoraLoggerMixin,viewsets.ModelViewSet):
    queryset=DetalleBitacora.objects.all()
    serializer_class=DetalleBitacoraSerializer
    permission_classes=[IsAuthenticated]
    
class DetalleBitacoraViewSet(BitacoraLoggerMixin,viewsets.ModelViewSet):
    queryset=DetalleBitacora.objects.all()
    serializer_class=DetalleBitacoraSerializer
    permission_classes=[IsAuthenticated]


from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db import transaction
import stripe
import logging

# Configuración de logger
logger = logging.getLogger(__name__)

from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
# views.py
from django.core.mail import send_mail
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.conf import settings
class SendEmailView(APIView):
    def post(self, request):
        subject = request.data.get('subject')
        message = request.data.get('message')

        if not subject or not message:
            return Response(
                {"error": "Subject and message are required."},
                status=status.HTTP_400_BAD_REQUEST
            )
        recipient_email = 'fredtaylorwolframio@gmail.com'

        try:
            print(f"Sending email to {recipient_email} with subject: {subject}")
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,  # Usar el correo configurado en settings.py
                [recipient_email],  # Lista de destinatarios
                fail_silently=False,  # Si hay un error, lanzará una excepción
            )
            print("Email sent successfully!")
            return Response({"message": "Email sent successfully!"}, status=status.HTTP_200_OK)
        except Exception as e:
            print(f"Error occurred: {e}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class PagoViewSet(BitacoraLoggerMixin, viewsets.ModelViewSet):
    queryset = Pago.objects.all()
    serializer_class = PagoSerializer

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def mis_comprobantes(self, request):
        """
        Devuelve todos los comprobantes de los pagos realizados por el usuario autenticado.
        """
        # Filtrar los pagos por el usuario autenticado
        pagos = Pago.objects.filter(usuario=request.user)

        # Serializar los datos de los pagos, incluyendo los comprobantes
        pagos_serializados = []
        for pago in pagos:
            pagos_serializados.append({
                'id': pago.id,
                'usuario_nombre': f"{pago.usuario.nombre} {pago.usuario.apellido_paterno}",
                'monto': str(pago.monto),  # Convertir el monto a string
                'fecha_pago': pago.fecha_pago.isoformat(),
                'metodo_pago': pago.metodo_pago,
                'referencia': pago.referencia,
                'comprobante': pago.comprobante.url if pago.comprobante else None,  # URL del comprobante
                'observaciones': pago.observaciones,
                'venta': pago.venta.id if pago.venta else None  # Asegurarse de que la venta esté asociada
            })

        return Response(pagos_serializados)


    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def crear_sesion_stripe(self, request):
        success_url = request.data.get('success_url')
        cancel_url = request.data.get('cancel_url')
        
        logger.info('🔐 Iniciando crear_sesion_stripe')
        logger.info(f"📥 URLs: success={success_url}, cancel={cancel_url}")

        # Validar las URLs
        if not all([success_url, cancel_url]):
            logger.error("❌ Faltan URLs")
            return Response(
                {"error": "Faltan success_url y/o cancel_url"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            with transaction.atomic():
                # Verificar si el usuario ya tiene un carrito "abierto"
                carrito = Carrito.objects.filter(user=request.user, estado='open').first()

                # Si no existe un carrito abierto, se devuelve un error
                if not carrito:
                    logger.error("❌ No existe un carrito abierto para este usuario.")
                    return Response(
                        {"error": "No existe un carrito abierto. Asegúrate de agregar productos al carrito."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                logger.info(f"🛒 Carrito obtenido: #{carrito.id} con {carrito.items.count()} items")

                # Verificar que el carrito tenga items
                if not carrito.items.exists():
                    logger.error("❌ Carrito vacío")
                    return Response(
                        {"error": "El carrito está vacío"}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Verificar stock de todos los items ANTES de crear la venta
                for item in carrito.items.select_related('producto').all():
                    if item.producto.stock < item.cantidad:
                        logger.error(
                            f"❌ Stock insuficiente: {item.producto.nombre} "
                            f"(disponible: {item.producto.stock}, requerido: {item.cantidad})"
                        )
                        return Response(
                            {
                                "error": f"Stock insuficiente para {item.producto.nombre}. "
                                         f"Disponible: {item.producto.stock}"
                            },
                            status=status.HTTP_400_BAD_REQUEST
                        )

                # Crear la venta
                venta = Venta.objects.create(
                    cliente=request.user,
                    carrito=carrito,
                    cantidad=carrito.items.count(),
                    precio_unitario = carrito.total / carrito.items.count() if carrito.items.count() else 0,
                    precio_total=carrito.total,
                    metodo_pago="tarjeta",
                    estado_venta="pendiente",
                )
                logger.info(f"✅ Venta creada: #{venta.id} - Total: ${venta.precio_total}")

                # Crear línea de items para Stripe
                line_items = []
                for item in carrito.items.select_related('producto').all():
                    line_items.append({
                        'price_data': {
                            'currency': 'BOB',
                            'product_data': {
                                'name': item.producto.nombre,
                            },
                            'unit_amount': int(item.producto.precio * 100),
                        },
                        'quantity': item.cantidad,
                    })
                
                logger.info(f"💳 Line items para Stripe: {len(line_items)} productos")

                # Crear sesión de Stripe
                session = stripe.checkout.Session.create(
                    payment_method_types=['card'],
                    line_items=line_items,
                    mode='payment',
                    success_url=success_url + '?session_id={CHECKOUT_SESSION_ID}',
                    cancel_url=cancel_url,
                    metadata={
                        'carrito_id': str(carrito.id),  # Enviar ID del carrito
                        'usuario_id': str(request.user.id),
                        'venta_id': str(venta.id)
                    }
                )

                logger.info(f"✅ Sesión Stripe creada: {session.id}")

                return Response({
                    'id': session.id,
                    'url': session.url,
                    'carrito_id': carrito.id,  # Devolver para referencia
                    'venta_id': venta.id
                })

        except Exception as e:
            logger.error(f"❌ Error al procesar sesión: {str(e)}", exc_info=True)
            return Response(
                {"error": str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


from .models import Promocion
from .serializers import PromocionSerializer

class PromocionViewSet(viewsets.ModelViewSet):
    queryset = Promocion.objects.all()
    serializer_class = PromocionSerializer
         


class VentaViewSet(viewsets.ModelViewSet):
    queryset = Venta.objects.all()
    serializer_class = VentaSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Si el usuario es administrador, devuelve todas las ventas
        if self.request.user.rol.nombre == 'Administrador':
            return Venta.objects.all()
        
        # Si el usuario no es administrador, devuelve solo sus propias ventas
        return Venta.objects.filter(cliente=self.request.user)
from .models import VentaProducto
from .serializers import VentaProductoSerializer

class VentaProductoViewSet(viewsets.ModelViewSet):
    queryset = VentaProducto.objects.all()
    serializer_class = VentaProductoSerializer
from .models import Producto
from .serializers import ProductoSerializer

class ProductoViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.all()
    serializer_class = ProductoSerializer



from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from .models import Carrito, CarritoItem, Producto
import logging

logger = logging.getLogger(__name__)

class CarritoViewSet(viewsets.ModelViewSet):
    queryset = Carrito.objects.all()
    serializer_class = CarritoSerializer

    def get_queryset(self):
        """Devuelve los carritos del usuario o el carrito anónimo."""
        if self.request.user.is_authenticated:
            return Carrito.objects.filter(user=self.request.user, estado='open')
        
        cart_token = self.request.query_params.get('cart_token')
        if cart_token:
            return Carrito.objects.filter(cart_token=cart_token, estado='open')
        
        return Carrito.objects.none()

    def get_or_create_cart(self, user):
        """Obtiene el carrito del usuario o crea uno si no existe."""
        carrito = Carrito.objects.filter(user=user, estado='open').first()

        # Si no existe un carrito abierto, lo creamos
        if not carrito:
            carrito = Carrito.objects.create(user=user, estado='open')
            logger.info(f"🛒 Carrito creado: #{carrito.id}")

        return carrito

    @action(detail=False, methods=['post'], url_path='add_item')
    def add_item(self, request):
        """Agregar un producto al carrito."""
        try:
            # Pasamos el usuario al método get_or_create_cart
            carrito = self.get_or_create_cart(request.user)
            logger.info(f"🛒 Carrito obtenido: #{carrito.id}")

            producto_id = request.data.get('producto')  # ID del producto a agregar
            cantidad = int(request.data.get('cantidad', 1))  # Cantidad a agregar

            if not producto_id:
                return Response(
                    {'error': 'El ID del producto es requerido'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if cantidad <= 0:
                return Response(
                    {'error': 'La cantidad debe ser mayor a 0'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            producto = get_object_or_404(Producto, id=producto_id)  # Obtén el producto

            # Verificar stock disponible
            if producto.stock < cantidad:
                return Response(
                    {'error': f'Stock insuficiente. Solo hay {producto.stock} unidades disponibles'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Usamos get_or_create para evitar duplicados, pero permitir la actualización de cantidad
            item, created = CarritoItem.objects.get_or_create(
                carrito=carrito,
                producto=producto,
                defaults={'cantidad': cantidad}
            )

            if not created:  # Si el item ya existía, solo actualizamos la cantidad
                nueva_cantidad = item.cantidad + cantidad
                if producto.stock < nueva_cantidad:
                    return Response(
                        {'error': f'Stock insuficiente. Solo hay {producto.stock} unidades'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                item.cantidad = nueva_cantidad
                item.save()
                logger.info(f"📦 Item actualizado: {producto.nombre} x{item.cantidad}")
            else:
                logger.info(f"📦 Item creado: {producto.nombre} x{cantidad}")

            # Retornamos los datos del carrito actualizado
            return Response({
                'message': 'Producto agregado al carrito exitosamente',
                'item': CarritoItemSerializer(item).data,
                'carrito': CarritoSerializer(carrito).data,
                'cart_token': carrito.cart_token
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            logger.error(f"❌ Error en add_item: {str(e)}", exc_info=True)
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='my_cart')
    def my_cart(self, request):
        """Obtener el carrito actual del usuario."""
        try:
            # Pasamos el usuario al método get_or_create_cart
            carrito = self.get_or_create_cart(request.user)
            logger.info(f"🛒 My cart - Retornando carrito #{carrito.id}")
            return Response(CarritoSerializer(carrito).data)
        except Exception as e:
            logger.error(f"❌ Error en my_cart: {str(e)}", exc_info=True)
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='remove_item')
    def remove_item(self, request):
        """
        Eliminar un ítem del carrito.
        URL: POST /api/carritos/remove_item/
        Body: { "producto": 1 }
        """
        try:
            # Pasamos el usuario al método get_or_create_cart
            carrito = self.get_or_create_cart(request.user)  # Aquí pasamos request.user
            producto_id = request.data.get('producto')
            
            if not producto_id:
                return Response(
                    {'error': 'El ID del producto es requerido'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            item = get_object_or_404(CarritoItem, carrito=carrito, producto_id=producto_id)
            item.delete()

            return Response({
                'message': 'Producto eliminado del carrito',
                'carrito': CarritoSerializer(carrito).data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='update_item')
    def update_item(self, request):
        """
        Actualizar la cantidad de un producto en el carrito.
        URL: POST /api/carritos/update_item/
        Body: { "producto": 1, "cantidad": 3 }
        """
        try:
            # Pasamos el usuario al método get_or_create_cart
            carrito = self.get_or_create_cart(request.user)  # Aquí pasamos request.user
            producto_id = request.data.get('producto')
            cantidad = int(request.data.get('cantidad', 1))
            
            if not producto_id:
                return Response(
                    {'error': 'El ID del producto es requerido'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if cantidad <= 0:
                return Response(
                    {'error': 'La cantidad debe ser mayor a 0'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            item = get_object_or_404(CarritoItem, carrito=carrito, producto_id=producto_id)
            
            # Verificar stock
            if item.producto.stock < cantidad:
                return Response(
                    {'error': f'Stock insuficiente. Solo hay {item.producto.stock} unidades disponibles'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            item.cantidad = cantidad
            item.save()

            return Response({
                'message': 'Cantidad actualizada',
                'item': CarritoItemSerializer(item).data,
                'carrito': CarritoSerializer(carrito).data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='my_cart')
    def my_cart(self, request):
        """
        Obtener el carrito actual del usuario.
        URL: GET /api/carritos/my_cart/
        Query params (opcional para anónimos): ?cart_token=xxx
        """
        try:
            carrito = self.get_or_create_cart(request.user)  # Aquí pasamos request.user
            return Response(CarritoSerializer(carrito).data)
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='vaciar_carrito')
    def vaciar_carrito(self, request):
        """
        Vaciar el carrito completamente.
        URL: POST /api/carritos/clear_cart/
        """
        try:
            # Pasamos el usuario al método get_or_create_cart
            carrito = self.get_or_create_cart(request.user)  # Aquí pasamos request.user
            carrito.items.all().delete()
            
            return Response({
                'message': 'Carrito vaciado exitosamente',
                'carrito': CarritoSerializer(carrito).data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )




class CarritoItemViewSet(viewsets.ModelViewSet):
    queryset = CarritoItem.objects.all()
    serializer_class = CarritoItemSerializer
    
    def get_queryset(self):
        """Filtrar items según el carrito del usuario."""
        if self.request.user.is_authenticated:
            return CarritoItem.objects.filter(carrito__user=self.request.user)
        cart_token = self.request.query_params.get('cart_token')
        if cart_token:
            return CarritoItem.objects.filter(carrito__cart_token=cart_token)
        return CarritoItem.objects.none()


from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework import viewsets
from .models import TipoGarantia
from .serializers import TipoGarantiaSerializer

class TipoGarantiaViewSet(viewsets.ViewSet):
    queryset = TipoGarantia.objects.all()  # Define tu queryset aquí
    serializer_class = TipoGarantiaSerializer
    
    # Crear un nuevo tipo de garantía
    @action(detail=False, methods=['post'], url_path='crear')
    def crear(self, request):
        serializer = TipoGarantiaSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()  # Guarda el nuevo tipo de garantía
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    # Obtener la lista de tipos de garantía
    @action(detail=False, methods=['get'], url_path='listar')
    def listar(self, request):
        queryset = TipoGarantia.objects.all()  # Obtiene todos los tipos de garantía
        serializer = TipoGarantiaSerializer(queryset, many=True)
        return Response(serializer.data)
    
    # Obtener un tipo de garantía específico
    @action(detail=True, methods=['get'], url_path='detalle')
    def detalle(self, request, pk=None):
        tipo_garantia = self.get_object()  # Obtiene el tipo de garantía por su ID
        serializer = TipoGarantiaSerializer(tipo_garantia)
        return Response(serializer.data)
    
    # Actualizar un tipo de garantía (completo)
    @action(detail=True, methods=['put'], url_path='actualizar')
    def actualizar(self, request, pk=None):
        tipo_garantia = self.get_object()  # Obtiene el tipo de garantía que se va a actualizar
        serializer = TipoGarantiaSerializer(tipo_garantia, data=request.data, partial=False)
        if serializer.is_valid():
            serializer.save()  # Guarda los cambios
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    # Actualizar un tipo de garantía (parcial)
    @action(detail=True, methods=['patch'], url_path='editar')
    def editar(self, request, pk=None):
        tipo_garantia = self.get_object()  # Obtiene el tipo de garantía a editar
        serializer = TipoGarantiaSerializer(tipo_garantia, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()  # Guarda los cambios parciales
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    # Eliminar un tipo de garantía
    @action(detail=True, methods=['delete'], url_path='eliminar')
    def eliminar(self, request, pk=None):
        tipo_garantia = self.get_object()  # Obtiene el tipo de garantía a eliminar
        tipo_garantia.delete()  # Elimina el tipo de garantía
        return Response(status=status.HTTP_204_NO_CONTENT)


from django.utils import timezone

# Webhook to handle Stripe events
# Webhook to handle Stripe events
from django.db import transaction
#$0.50 USD
@csrf_exempt
def stripe_webhook(request):
    payload = request.body
    sig_header = request.META.get('HTTP_STRIPE_SIGNATURE')

    if not sig_header:
        logger.error("❌ No se recibió firma de Stripe")
        return HttpResponse(status=400)

    event = None

    try:
        event = stripe.Webhook.construct_event(
            payload, 
            sig_header, 
            settings.STRIPE_WEBHOOK_SECRET
        )
    except ValueError as e:
        logger.error(f"❌ Payload inválido: {str(e)}")
        return HttpResponse(status=400)
    except stripe.error.SignatureVerificationError as e:
        logger.error(f"❌ Firma inválida: {str(e)}")
        return HttpResponse(status=400)

    if event['type'] == 'checkout.session.completed':
        session = event['data']['object']
        metadata = session.get('metadata', {})
        
        venta_id = metadata.get('venta_id')
        usuario_id = metadata.get('usuario_id')
        carrito_id = metadata.get('carrito_id')  # ✅ Obtener carrito_id

        logger.info(
            f"🔔 Webhook recibido - Venta: {venta_id}, "
            f"Usuario: {usuario_id}, Carrito: {carrito_id}"
        )

        if not all([venta_id, usuario_id, carrito_id]):
            logger.error(f"❌ Metadata incompleta: {metadata}")
            return HttpResponse(status=400)

        try:
            usuario = Usuario.objects.get(id=usuario_id)
            monto = session['amount_total'] / 100
            referencia = session.get('payment_intent') or session.get('id')

            logger.info(f"💰 Procesando pago - ${monto} - Ref: {referencia}")

            # Evitar duplicados
            if Pago.objects.filter(referencia=referencia, usuario=usuario).exists():
                logger.warning(f"⚠️ Pago duplicado: {referencia}")
                return HttpResponse(status=200)

            # Obtener venta y carrito
            venta = Venta.objects.select_related('carrito').get(id=venta_id)
            
            # ✅ VERIFICAR que el carrito de la venta coincida
            if str(venta.carrito.id) != str(carrito_id):
                logger.error(
                    f"❌ Mismatch de carrito: venta tiene {venta.carrito.id}, "
                    f"metadata tiene {carrito_id}"
                )
                return HttpResponse(status=400)

            with transaction.atomic():
                # Obtener carrito con items
                carrito = Carrito.objects.prefetch_related(
                    'items__producto'
                ).get(id=carrito_id)
                
                items_count = carrito.items.count()
                logger.info(f"🛒 Procesando carrito #{carrito.id} con {items_count} items")

                if items_count == 0:
                    logger.error(f"❌ Carrito #{carrito.id} está vacío")
                    return HttpResponse(status=400)

                # Procesar items
                items_procesados = 0
                for item in carrito.items.select_related('producto').all():
                    # Verificar stock
                    if item.producto.stock < item.cantidad:
                        logger.error(
                            f"❌ Stock insuficiente: {item.producto.nombre} "
                            f"(disponible: {item.producto.stock}, requerido: {item.cantidad})"
                        )
                        return HttpResponse(
                            content=f"Stock insuficiente para {item.producto.nombre}",
                            status=400
                        )
                    
                    # Crear VentaProducto
                    VentaProducto.objects.create(
                        venta=venta,
                        producto=item.producto,
                        cantidad=item.cantidad,
                        precio_unitario=item.precio_unitario
                    )
                    
                    # Actualizar stock
                    item.producto.stock -= item.cantidad
                    item.producto.save(update_fields=['stock'])
                    
                    items_procesados += 1
                    logger.info(
                        f"✅ Item procesado: {item.producto.nombre} x{item.cantidad} "
                        f"(Stock restante: {item.producto.stock})"
                    )

                logger.info(f"✅ {items_procesados} items procesados exitosamente")

                # Crear pago
                pago = Pago.objects.create(
                    usuario=usuario,
                    monto=monto,
                    metodo_pago='tarjeta',
                    referencia=referencia,
                    fecha_pago=timezone.now(),
                    venta=venta
                )
                logger.info(f"💳 Pago creado: #{pago.id} - ${monto}")

                # Actualizar estados
                venta.estado_venta = 'completada'
                venta.save(update_fields=['estado_venta'])
                
                carrito.estado = 'ordered'
                carrito.save(update_fields=['estado'])
                
                logger.info(f"✅ Venta #{venta.id} completada, Carrito #{carrito.id} cerrado")

            logger.info(f"🎉 Proceso completado exitosamente para venta #{venta.id}")
            return HttpResponse(status=200)

        except Exception as e:
            logger.error(f"❌ Error inesperado en webhook: {str(e)}", exc_info=True)
            return HttpResponse(status=500)

    return HttpResponse(status=200)
from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework import status
from .models import ReporteEstatico
from .serializers import ReporteEstaticoSerializer
import openpyxl

from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
import openpyxl
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.shortcuts import render
from .models import Venta, VentaProducto

class ReporteEstaticoViewSet(viewsets.ModelViewSet):
    queryset = ReporteEstatico.objects.all()
    serializer_class = ReporteEstaticoSerializer

    @action(detail=False, methods=['get'])
    def reporte_ventas(self, request):
        # Obtener el formato desde los parámetros de la URL (por defecto Excel)
        formato = request.GET.get('formato', 'excel')

        # Filtramos las ventas
        ventas = Venta.objects.all()

        # Crear el reporte en el formato solicitado
        if formato == 'pdf':
            return self.generar_reporte_pdf(ventas,request)
        elif formato == 'html':
            return self.generar_reporte_html(ventas,request)
        else:
            return self.generar_reporte_excel(ventas,request)

    def generar_reporte_excel(self, ventas,request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Ventas"

        ws['A1'] = 'Cliente'
        ws['B1'] = 'Producto'
        ws['C1'] = 'Cantidad'
        ws['D1'] = 'Precio Unitario'
        ws['E1'] = 'Total'
        ws['F1'] = 'Fecha de Venta'

        row = 2  # Comenzamos en la segunda fila para los datos
        for venta in ventas:
            venta_productos = VentaProducto.objects.filter(venta=venta)
            for vp in venta_productos:
                ws[f'A{row}'] = venta.cliente.username
                ws[f'B{row}'] = vp.producto.nombre
                ws[f'C{row}'] = vp.cantidad
                ws[f'D{row}'] = vp.precio_unitario
                ws[f'E{row}'] = vp.cantidad * vp.precio_unitario
                ws[f'F{row}'] = venta.fecha
                row += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Reporte_Ventas.xlsx'
        wb.save(response)
        return response

    def generar_reporte_pdf(self, ventas,request):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=Reporte_Ventas.pdf'

        p = canvas.Canvas(response, pagesize=letter)
        p.drawString(100, 750, "Reporte de Ventas")

        # Escribir encabezados
        p.drawString(100, 730, 'Cliente')
        p.drawString(200, 730, 'Producto')
        p.drawString(300, 730, 'Cantidad')
        p.drawString(400, 730, 'Precio Unitario')
        p.drawString(500, 730, 'Total')
        p.drawString(600, 730, 'Fecha de Venta')

        y_position = 710  # Empezamos a escribir debajo de los encabezados
        for venta in ventas:
            venta_productos = VentaProducto.objects.filter(venta=venta)
            for vp in venta_productos:
                p.drawString(100, y_position, venta.cliente.username)
                p.drawString(200, y_position, vp.producto.nombre)
                p.drawString(300, y_position, str(vp.cantidad))
                p.drawString(400, y_position, str(vp.precio_unitario))
                p.drawString(500, y_position, str(vp.cantidad * vp.precio_unitario))
                p.drawString(600, y_position, str(venta.fecha))
                y_position -= 20

        p.showPage()
        p.save()
        return response

    def generar_reporte_html(self, ventas, request):
        html_content = """
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <title>Reporte de Ventas</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; }
            table { width: 100%; border-collapse: collapse; margin-top: 20px; }
            th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
            th { background-color: #4CAF50; color: white; }
            tr:nth-child(even) { background-color: #f2f2f2; }
            h1 { color: #333; }
        </style>
    </head>
    <body>
        <h1>Reporte de Ventas</h1>
        <table>
            <thead>
                <tr>
                    <th>Cliente</th>
                    <th>Producto</th>
                    <th>Cantidad</th>
                    <th>Precio Unitario</th>
                    <th>Total</th>
                    <th>Fecha de Venta</th>
                </tr>
            </thead>
            <tbody>
    """
    
        for venta in ventas:
            venta_productos = VentaProducto.objects.filter(venta=venta)
            for vp in venta_productos:
                total = vp.cantidad * vp.precio_unitario
                html_content += f"""
                    <tr>
                        <td>{venta.cliente.username}</td>
                    <td>{vp.producto.nombre}</td>
                    <td>{vp.cantidad}</td>
                    <td>${vp.precio_unitario}</td>
                    <td>${total}</td>
                    <td>{venta.fecha}</td>
                </tr>
            """
    
        html_content += """
            </tbody>
        </table>
        </body>
        </html>
        """
    
        return HttpResponse(html_content)

    
    # Listar todos los reportes estáticos
    def listar(self, request):
        queryset = ReporteEstatico.objects.all()
        serializer = ReporteEstaticoSerializer(queryset, many=True)
        return Response(serializer.data)
    
    # Detalle de un reporte estático específico
    def detalle(self, request, pk=None):
        try:
            reporte = ReporteEstatico.objects.get(pk=pk)
        except ReporteEstatico.DoesNotExist:
            return Response({"detail": "Reporte no encontrado"}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = ReporteEstaticoSerializer(reporte)
        return Response(serializer.data)
    
    # Crear un nuevo reporte estático
    def crear(self, request):
        serializer = ReporteEstaticoSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    # Actualizar un reporte estático (Reemplaza todo)
    def actualizar(self, request, pk=None):
        try:
            reporte = ReporteEstatico.objects.get(pk=pk)
        except ReporteEstatico.DoesNotExist:
            return Response({"detail": "Reporte no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        serializer = ReporteEstaticoSerializer(reporte, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    # Editar un reporte estático (Actualización parcial)
    def editar(self, request, pk=None):
        try:
            reporte = ReporteEstatico.objects.get(pk=pk)
        except ReporteEstatico.DoesNotExist:
            return Response({"detail": "Reporte no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        serializer = ReporteEstaticoSerializer(reporte, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    # Eliminar un reporte estático
    def eliminar(self, request, pk=None):
        try:
            reporte = ReporteEstatico.objects.get(pk=pk)
        except ReporteEstatico.DoesNotExist:
            return Response({"detail": "Reporte no encontrado"}, status=status.HTTP_404_NOT_FOUND)
        
        reporte.delete()
        return Response({"detail": "Reporte eliminado"}, status=status.HTTP_204_NO_CONTENT)
from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework import status
from .models import ReporteDinamico
from .serializers import ReporteDinamicoSerializer

from rest_framework import viewsets, status
from rest_framework.response import Response
from .models import ReporteDinamico, Venta, Producto  # Importar otras tablas según sea necesario
from .serializers import ReporteDinamicoSerializer
from django.http import HttpResponse
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

class ReporteDinamicoViewSet(viewsets.ModelViewSet):
    queryset = ReporteDinamico.objects.all()
    serializer_class = ReporteDinamicoSerializer

    def generar_repomrte_excel(self, tabla, campos, request):
        # Filtrar datos según la tabla y los campos seleccionados
        if tabla == 'venta':
            ventas = Venta.objects.all()
            # Lógica de generación de reporte en Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte de Ventas"
            for idx, campo in enumerate(campos, start=1):
                ws[f'{chr(64 + idx)}1'] = campo  # Asignamos los nombres de los campos en la primera fila

            row = 2
            for venta in ventas:
                for idx, campo in enumerate(campos, start=1):
                    valor = getattr(venta, campo, '')  # Extraemos el valor del campo
                    ws[f'{chr(64 + idx)}{row}'] = valor
                row += 1

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=Reporte_Ventas.xlsx'
            wb.save(response)
            return response

    def generar_reporte_pdf(self, tabla, campos, request):
        # Filtrar datos según la tabla y los campos seleccionados
        if tabla == 'venta':
            ventas = Venta.objects.all()
            # Lógica de generación de reporte en PDF
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename=Reporte_Ventas.pdf'
            p = canvas.Canvas(response, pagesize=letter)
            y_position = 750

            # Escribir encabezados
            for idx, campo in enumerate(campos):
                p.drawString(100 + (idx * 100), y_position, campo)

            y_position -= 20
            for venta in ventas:
                for idx, campo in enumerate(campos):
                    valor = getattr(venta, campo, '')  # Extraemos el valor del campo
                    p.drawString(100 + (idx * 100), y_position, str(valor))
                y_position -= 20

            p.showPage()
            p.save()
            return response

    def generar_reporte_html(self, tabla, campos, request):
        # Filtrar datos según la tabla y los campos seleccionados
        if tabla == 'venta':
            ventas = Venta.objects.all()
            html_content = """
            <!DOCTYPE html>
            <html lang="es">
            <head>
                <meta charset="UTF-8">
                <title>Reporte de Ventas</title>
                <style>
                    body { font-family: Arial, sans-serif; margin: 20px; }
                    table { width: 100%; border-collapse: collapse; margin-top: 20px; }
                    th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                    th { background-color: #4CAF50; color: white; }
                    tr:nth-child(even) { background-color: #f2f2f2; }
                </style>
            </head>
            <body>
                <h1>Reporte de Ventas</h1>
                <table>
                    <thead>
                        <tr>
            """
            for campo in campos:
                html_content += f"<th>{campo}</th>"

            html_content += """
                    </tr>
                </thead>
                <tbody>
            """

            for venta in ventas:
                html_content += "<tr>"
                for campo in campos:
                    valor = getattr(venta, campo, '')  # Extraemos el valor del campo
                    html_content += f"<td>{valor}</td>"
                html_content += "</tr>"

            html_content += """
                </tbody>
            </table>
            </body>
            </html>
            """

            return HttpResponse(html_content)

    @action(detail=False, methods=['get'])
    def generar_reporte(self, request):
        # Obtenemos los parámetros de la URL: tabla y campos
        tabla = request.GET.get('tabla')
        campos = request.GET.getlist('campos')  # Lista de campos seleccionados
        formato = request.GET.get('formato', 'excel')

        if formato == 'pdf':
            return self.generar_reporte_pdf(tabla, campos, request)
        elif formato == 'html':
            return self.generar_reporte_html(tabla, campos, request)
        else:
            return self.generar_reporte_excel(tabla, campos, request)
  
    # Listar todos los reportes dinámicos
    def listar(self, request):
        queryset = ReporteDinamico.objects.all()
        serializer = ReporteDinamicoSerializer(queryset, many=True)
        return Response(serializer.data)
    
    # Detalle de un reporte dinámico específico
    def detalle(self, request, pk=None):
        try:
            reporte = ReporteDinamico.objects.get(pk=pk)
        except ReporteDinamico.DoesNotExist:
            return Response({"detail": "Reporte no encontrado"}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = ReporteDinamicoSerializer(reporte)
        return Response(serializer.data)
    
    # Crear un nuevo reporte dinámico
    def crear(self, request):
        serializer = ReporteDinamicoSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    # Actualizar un reporte dinámico (Reemplaza todo)
    def actualizar(self, request, pk=None):
        try:
            reporte = ReporteDinamico.objects.get(pk=pk)
        except ReporteDinamico.DoesNotExist:
            return Response({"detail": "Reporte no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        serializer = ReporteDinamicoSerializer(reporte, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    # Editar un reporte dinámico (Actualización parcial)
    def editar(self, request, pk=None):
        try:
            reporte = ReporteDinamico.objects.get(pk=pk)
        except ReporteDinamico.DoesNotExist:
            return Response({"detail": "Reporte no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        serializer = ReporteDinamicoSerializer(reporte, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    # Eliminar un reporte dinámico
    def eliminar(self, request, pk=None):
        try:
            reporte = ReporteDinamico.objects.get(pk=pk)
        except ReporteDinamico.DoesNotExist:
            return Response({"detail": "Reporte no encontrado"}, status=status.HTTP_404_NOT_FOUND)
        
        reporte.delete()
        return Response({"detail": "Reporte eliminado"}, status=status.HTTP_204_NO_CONTENT)

from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework import status
from .models import ReporteIA
from .serializers import ReporteIASerializer

class ReporteIAViewSet(viewsets.ViewSet):
    queryset = ReporteIA.objects.all()
    serializer_class = ReporteIASerializer    
    
    # Listar todos los reportes de IA
    def listar(self, request):
        queryset = ReporteIA.objects.all()
        serializer = ReporteIASerializer(queryset, many=True)
        return Response(serializer.data)
    
    # Detalle de un reporte de IA específico
    def detalle(self, request, pk=None):
        try:
            reporte = ReporteIA.objects.get(pk=pk)
        except ReporteIA.DoesNotExist:
            return Response({"detail": "Reporte no encontrado"}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = ReporteIASerializer(reporte)
        return Response(serializer.data)
    
    # Crear un nuevo reporte de IA
    def crear(self, request):
        serializer = ReporteIASerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    # Actualizar un reporte de IA (Reemplaza todo)
    def actualizar(self, request, pk=None):
        try:
            reporte = ReporteIA.objects.get(pk=pk)
        except ReporteIA.DoesNotExist:
            return Response({"detail": "Reporte no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        serializer = ReporteIASerializer(reporte, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    # Editar un reporte de IA (Actualización parcial)
    def editar(self, request, pk=None):
        try:
            reporte = ReporteIA.objects.get(pk=pk)
        except ReporteIA.DoesNotExist:
            return Response({"detail": "Reporte no encontrado"}, status=status.HTTP_404_NOT_FOUND)

        serializer = ReporteIASerializer(reporte, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    # Eliminar un reporte de IA
    def eliminar(self, request, pk=None):
        try:
            reporte = ReporteIA.objects.get(pk=pk)
        except ReporteIA.DoesNotExist:
            return Response({"detail": "Reporte no encontrado"}, status=status.HTTP_404_NOT_FOUND)
        
        reporte.delete()
        return Response({"detail": "Reporte eliminado"}, status=status.HTTP_204_NO_CONTENT)
# views.py
from rest_framework.views import APIView
from rest_framework.response import Response
from django.db.models import Sum
from .models import Venta, VentaProducto

class VentasHistoricasAPIView(APIView):
    def get(self, request, *args, **kwargs):
        # Obtener ventas agrupadas por producto
        ventas = VentaProducto.objects.values('producto_id') \
            .annotate(total_ventas=Sum('precio_unitario')) \
            .order_by('producto_id')
        return Response(ventas)

import joblib
import pandas as pd
from sklearn.preprocessing import LabelEncoder
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from app.models import Venta, Feriado
import os
from django.conf import settings
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error

# Obtener la ruta completa al modelo
model_path = os.path.join(settings.BASE_DIR, 'model_training', 'random_forest_sales_model.pkl')

# Cargar el modelo si ya existe
if os.path.exists(model_path):
    model = joblib.load(model_path)
else:
    model = None  # Si el modelo no existe, lo dejamos como None

class VentaModelView(APIView):

    def post(self, request, *args, **kwargs):
        action = request.data.get('action')  # Determinar si es entrenamiento o predicción

        if action == 'train':
            return self.entrenar_modelo(request)
        elif action == 'predict':
            return self.hacer_prediccion(request)
        else:
            return Response({'error': 'Acción no válida, use "train" o "predict".'}, status=status.HTTP_400_BAD_REQUEST)

    def hacer_prediccion(self, request):
        try:
            # Extraer los datos de la solicitud (esperamos que sean datos JSON)
            cantidad = request.data['cantidad']
            precio_unitario = request.data['precio_unitario']
            promocion = request.data['promocion']
            metodo_pago = request.data['metodo_pago']
            dia_de_la_semana = request.data['dia_de_la_semana']
            mes_del_ano = request.data['mes_del_ano']
            es_feriado = request.data['es_feriado']

            # Codificar 'metodo_pago' con LabelEncoder (suponiendo que ya lo entrenaste como en el código anterior)
            le = LabelEncoder()
            metodo_pago_codificado = le.fit_transform([metodo_pago])[0]  # Asegúrate de que 'metodo_pago' tenga el mismo formato

            # Crear un DataFrame con los datos proporcionados
            nueva_venta_df = pd.DataFrame([{
                'cantidad': cantidad,
                'precio_unitario': precio_unitario,
                'promocion': promocion,
                'metodo_pago': metodo_pago_codificado,
                'dia_de_la_semana': dia_de_la_semana,
                'mes_del_ano': mes_del_ano,
                'es_feriado': es_feriado
            }])

            # Realizar la predicción
            prediccion = model.predict(nueva_venta_df)

            # Retornar la predicción como respuesta
            return Response({'precio_total_predicho': prediccion[0]}, status=status.HTTP_200_OK)
        
        except KeyError as e:
            # Si falta alguno de los parámetros, retornar un error
            return Response({'error': f'Faltan parámetros en la solicitud: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            # Manejo de errores genéricos
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def entrenar_modelo(self, request):
        try:
            # Obtener las ventas
            ventas = Venta.objects.all()

            # Crear un DataFrame con los datos de las ventas
            data = []
            for venta in ventas:
                # Convertir fecha de la venta a características útiles
                dia_de_la_semana = venta.fecha.weekday()
                mes_del_ano = venta.fecha.month
                es_feriado = Feriado.objects.filter(fecha=venta.fecha).exists()

                # Agregar los datos a la lista
                data.append({
                    'fecha': venta.fecha,
                    'cantidad': venta.cantidad,
                    'precio_unitario': venta.precio_unitario,
                    'precio_total': venta.precio_total,
                    'promocion': 1 if venta.promocion else 0,
                    'metodo_pago': venta.metodo_pago,
                    'dia_de_la_semana': dia_de_la_semana,
                    'mes_del_ano': mes_del_ano,
                    'es_feriado': es_feriado
                })

            # Crear un DataFrame de Pandas
            df = pd.DataFrame(data)
            le = LabelEncoder()
            df['metodo_pago'] = le.fit_transform(df['metodo_pago'])

            # Dividir las características (X) y el objetivo (y)
            X = df[['cantidad', 'precio_unitario', 'promocion', 'metodo_pago', 'dia_de_la_semana', 'mes_del_ano', 'es_feriado']]
            y = df['precio_total']

            # Dividir los datos en conjunto de entrenamiento y prueba
            X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

            # Inicializar el modelo RandomForestRegressor
            model = RandomForestRegressor(n_estimators=100, random_state=42)

            # Entrenar el modelo
            model.fit(X_train, y_train)

            # Evaluar el modelo
            y_pred = model.predict(X_test)
            mse = mean_squared_error(y_test, y_pred)
            print(f"Error cuadrático medio (MSE): {mse}")

            # Guardar el modelo entrenado en un archivo
            model_path = os.path.join(settings.BASE_DIR, 'model_training', 'random_forest_sales_model.pkl')
            joblib.dump(model, model_path)
            print(f"Modelo guardado en: {model_path}")

            return Response({'message': 'Modelo entrenado exitosamente.', 'mse': mse}, status=status.HTTP_200_OK)

        except Exception as e:
            # Manejo de errores durante el entrenamiento
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from google.cloud import dialogflow_v2 as dialogflow
from .models import Producto, Carrito, CarritoItem  # Asegúrate de importar tus modelos aquí
from django.db.models import Q

class DialogflowView(APIView):
    def post(self, request, *args, **kwargs):
        # Extraer la transcripción de la solicitud
        text = request.data.get("text")
        print(f"Texto recibido: {text}")  # Debugging

        if not text:
            return Response({"error": "El texto es requerido"}, status=status.HTTP_400_BAD_REQUEST)

        # Configurar la sesión de Dialogflow
        project_id = "proyecto2-2025"  # Cambia esto por tu ID de proyecto de Dialogflow
        session_id = "random-session-id"  # Puedes generar una sesión única
        session_client = dialogflow.SessionsClient()
        session = session_client.session_path(project_id, session_id)

        text_input = dialogflow.TextInput(text=text, language_code="es")
        query_input = dialogflow.QueryInput(text=text_input)

        try:
            # Detectar la intención en Dialogflow
            response = session_client.detect_intent(session=session, query_input=query_input)
            intent = response.query_result.intent.display_name
            fulfillment_text = response.query_result.fulfillment_text  # Respuesta de Dialogflow
            print(f"Intención detectada: {intent}")  # Debugging
            print(f"Fulfillment text: {fulfillment_text}")  # Debugging

            # Obtener los parámetros del intent (producto, marca, modelo)
            producto = response.query_result.parameters.get("producto")
            marca = response.query_result.parameters.get("marca")
            print(f"Parámetros detectados - Producto:{producto}, Marca: {marca}")  # Debugging

            # Limpiar los valores de producto y marca (eliminar espacios extras y asegurar que marca sea una cadena)
            marca = marca[0] 
            print(f"Marca limpia: {marca}")  # Debugging

            if intent == "AddProductToCart" and producto:
                # Búsqueda más flexible del producto por nombre, marca y modelo
                try:
                    # Asegúrate de que 'producto' no esté vacío y sea un valor válido para la búsqueda
                    if not producto:
                        return Response({"error": "El producto no está especificado correctamente."}, status=status.HTTP_400_BAD_REQUEST)
                    
                    productos = Producto.objects.filter(
                        Q(nombre__icontains=producto) & Q(marca__icontains=marca)
                    )

                    print(f"Productos encontrados: {productos}")  # Debugging

                    if productos.exists():
                        # Elegir el primer producto encontrado o uno basado en otro criterio
                        product = productos.first()
                        print(f"Producto encontrado: {product}")  # Debugging

                        # Obtener el carrito del usuario y agregar el producto al carrito
                        cart_token = request.data.get("cart_token")

                        if not cart_token:
                            # Si no se proporciona cart_token, crear uno nuevo
                            cart_token = get_random_string(48)  # Generar un nuevo token de carrito
                            print(f"Nuevo token de carrito generado: {cart_token}")

                        try:
                            # Buscar el carrito utilizando el token
                            carrito = Carrito.objects.get(cart_token=cart_token, estado="open")
                        except Carrito.DoesNotExist:
                            # Si el carrito no existe, crear uno nuevo
                            if request.user.is_authenticated:
                                carrito = Carrito.objects.create(cart_token=cart_token, estado="open", user=request.user)
                                print(f"Carrito creado: ")  # Debugging


                            

                        print(f"Carrito encontrado: {carrito}")  # Debugging
                        
                        # Verificar si el producto ya existe en el carrito
                        carrito_item, created = CarritoItem.objects.get_or_create(
                            carrito=carrito,
                            producto=product,
                            defaults={"cantidad": 1},
                        )

                        if not created:
                            carrito_item.cantidad += 1
                            carrito_item.save()
                            print(f"Cantidad del producto en el carrito actualizada: {carrito_item.cantidad}")  # Debugging

                        return Response({
                            "message": f"Producto {product.nombre} agregado al carrito.",
                            "fulfillment_text": fulfillment_text
                        })

                    else:
                        print(f"Producto no encontrado en el catálogo con los parámetros dados.")  # Debugging
                        return Response({"error": "Producto no encontrado en el catálogo"}, status=status.HTTP_404_NOT_FOUND)

                except Producto.DoesNotExist:
                    print(f"Error al buscar el producto.")  # Debugging
                    return Response({"error": "Producto no encontrado en el catálogo"}, status=status.HTTP_404_NOT_FOUND)

            return Response({"message": fulfillment_text})

        except Exception as e:
            print(f"Error al procesar la solicitud: {str(e)}")  # Debugging
            return Response({"error": f"Error al procesar la solicitud: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)





